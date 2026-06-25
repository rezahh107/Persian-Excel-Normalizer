"""
normalize_excel.py — Production-grade Persian text normaliser for Excel files.

Architecture
------------
* Pure-function core — normalize_text() and normalize_worksheet() are
  side-effect-free and trivially unit-testable.
* CLI is a thin adapter — main() only parses arguments and delegates.
* Structured logging on a dedicated "excel_normalizer.core" logger;
  the root logger is never touched.
* Optional progress / cancellation callbacks are keyword-only with None
  defaults so the public API remains fully backwards-compatible.
* pathlib is used throughout for OS-agnostic path handling.

Performance notes
-----------------
* normalize_text() uses str.translate() with two pre-built tables
  (_TABLE_REPLACE_ZWNJ / _TABLE_KEEP_ZWNJ) constructed at import time.
  A single translate() call replaces all Arabic→Persian substitutions,
  invisible-char deletions and NBSP normalisation in one O(n) pass,
  instead of the previous ~12 sequential .replace() calls.
* For workbooks larger than ~200 MB, openpyxl loads the entire file into
  RAM. If memory is a concern, consider splitting large workbooks before
  processing or migrating to a streaming library such as xlrd/xlwt.

Formula-safety
--------------
normalize_worksheet() skips cells whose value starts with '=' so that
formula strings (e.g. =SUM(A1:A10)) are never touched.
"""

from __future__ import annotations

import argparse
import logging
import re
import shutil
import sys
from pathlib import Path
from typing import Callable, Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Module-level logger — never pollutes the root logger
# ---------------------------------------------------------------------------

log: logging.Logger = logging.getLogger("excel_normalizer.core")

# ---------------------------------------------------------------------------
# Pre-built translate tables (constructed once at import time)
# ---------------------------------------------------------------------------
# Using str.translate() with a single combined table replaces 12+ sequential
# .replace() calls with one O(n) pass over the string.

_BASE_TABLE: dict[str, str | None] = {
    # ── Arabic → Persian substitutions ────────────────────────────────────
    "\u064a": "\u06cc",   # Arabic yeh        → Persian yeh
    "\u0643": "\u06a9",   # Arabic kaf         → Persian kaf
    "\u0649": "\u06cc",   # Alef maqsura       → Persian yeh
    "\u0629": "\u0647",   # Taa marbuta        → Persian heh
    "\u0623": "\u0627",   # Alef hamza above   → bare alef
    "\u0625": "\u0627",   # Alef hamza below   → bare alef
    "\u0624": "\u0648",   # Waw with hamza     → bare waw
    "\u0626": "\u06cc",   # Yeh with hamza     → Persian yeh
    # ── Invisible control chars → delete (None) ───────────────────────────
    "\u200d": None,        # ZWJ
    "\u200e": None,        # LRM
    "\u200f": None,        # RLM
    "\u202a": None,        # LRE
    "\u202b": None,        # RLE
    "\u202c": None,        # PDF
    "\u2066": None,        # LRI
    "\u2067": None,        # RLI
    "\u2068": None,        # FSI
    "\u2069": None,        # PDI
    "\xad":   None,        # SHY (soft hyphen)
    # ── NBSP → regular space ──────────────────────────────────────────────
    "\xa0": " ",
}

# ZWNJ (U+200C) is handled conditionally — two tables, selected per call.
_TABLE_REPLACE_ZWNJ = str.maketrans({**_BASE_TABLE, "\u200c": " "})
_TABLE_KEEP_ZWNJ    = str.maketrans(_BASE_TABLE)

_MULTI_SPACE_RE: re.Pattern[str] = re.compile(r"[ \t]+")


# ---------------------------------------------------------------------------
# Pure normalisation functions  — DO NOT ALTER SEMANTICS
# ---------------------------------------------------------------------------


def normalize_text(text: str, *, replace_zwnj: bool = True) -> str:
    """Normalise a single Persian string.

    Uses a pre-built str.translate() table for a single O(n) pass
    instead of multiple sequential .replace() calls.

    Args:
        text: Raw string value to normalise.
        replace_zwnj: When *True* (default) U+200C is replaced with a space.

    Returns:
        Normalised string.  Non-string inputs are returned unchanged.
    """
    if not isinstance(text, str):
        return text  # type: ignore[return-value]

    table = _TABLE_REPLACE_ZWNJ if replace_zwnj else _TABLE_KEEP_ZWNJ
    text = text.translate(table)
    return _MULTI_SPACE_RE.sub(" ", text).strip()


def normalize_worksheet(
    sheet: Worksheet,
    *,
    replace_zwnj: bool = True,
    cancel_check: Optional[Callable[[], bool]] = None,
    row_callback: Optional[Callable[[int], None]] = None,
) -> int:
    """Normalise every eligible string cell in *sheet* in-place.

    Formula cells (value starts with '=') are intentionally skipped to
    prevent corruption of formula strings.

    Args:
        sheet: Writable openpyxl Worksheet.
        replace_zwnj: Forwarded to normalize_text().
        cancel_check: Optional callable; returns True to request cancellation.
            Checked once per row.  Output written so far is preserved.
        row_callback: Optional callable invoked after each row with the number
            of cells in that row.  Used for fine-grained GUI progress.

    Returns:
        Number of cells whose value was actually changed.
    """
    changed = 0
    for row in sheet.iter_rows():
        if cancel_check and cancel_check():
            log.info("Cancellation requested — stopping worksheet iteration.")
            break
        cells_in_row = 0
        for cell in row:
            cells_in_row += 1
            if not isinstance(cell.value, str):
                continue
            # ── Formula guard (fix: مشکل ۳) ───────────────────────────────
            # Cells whose value begins with '=' are formula cells.
            # Normalising them would corrupt the formula syntax.
            if cell.value.startswith("="):
                continue
            normalised = normalize_text(cell.value, replace_zwnj=replace_zwnj)
            if normalised != cell.value:
                cell.value = normalised
                changed += 1
        if row_callback is not None:
            row_callback(cells_in_row)
    return changed


# ---------------------------------------------------------------------------
# Workbook-level orchestration
# ---------------------------------------------------------------------------


def _estimate_total_cells(workbook: openpyxl.Workbook) -> int:
    """Return a conservative cell-count estimate across all sheets.

    Note: max_row × max_column over-counts sparse sheets.  The result is
    used only for progress-bar scaling, not for correctness.
    """
    total = 0
    for ws in workbook.worksheets:
        total += (ws.max_row or 0) * (ws.max_column or 0)
    return max(total, 1)


def _atomic_save(workbook: openpyxl.Workbook, destination: Path) -> None:
    """Save *workbook* to *destination* atomically via a temporary file.

    Writes to ``<destination>.tmp~`` first, then renames.  If the save
    fails the partially-written temp file is cleaned up and the original
    destination (if any) is left untouched.
    """
    tmp = destination.with_suffix(".tmp~")
    try:
        workbook.save(tmp)
        shutil.move(str(tmp), destination)
    except Exception:
        if tmp.exists():
            tmp.unlink(missing_ok=True)
        raise


def normalize_workbook(
    source: Path,
    destination: Path,
    *,
    replace_zwnj: bool = True,
    cancel_check: Optional[Callable[[], bool]] = None,
    on_sheet_start: Optional[Callable[[int, int, str], None]] = None,
    on_progress: Optional[Callable[[int, int], None]] = None,
) -> None:
    """Load *source*, normalise all worksheets, and save to *destination*.

    Memory note: openpyxl loads the entire workbook into RAM.  For files
    larger than ~200 MB this may be significant.  Consider splitting large
    workbooks before processing if memory is constrained.

    Args:
        source: Path to the input .xlsx / .xlsm workbook.
        destination: Where to write the normalised workbook.
        replace_zwnj: Forwarded to normalize_worksheet().
        cancel_check: Optional cancellation predicate.
        on_sheet_start: Optional callback ``(sheet_index, total, sheet_name)``.
        on_progress: Optional callback ``(processed_cells, total_cells)``
            for cell-level progress reporting.

    Raises:
        FileNotFoundError: If *source* does not exist.
        ValueError: If *source* has an unsupported extension.
        OSError: Propagated from openpyxl on read / write failure.
    """
    if not source.exists():
        raise FileNotFoundError(f"Input file not found: {source}")
    if source.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(
            f"Unsupported extension '{source.suffix}'. Expected .xlsx or .xlsm."
        )

    log.info("Loading workbook: %s", source)
    workbook = openpyxl.load_workbook(source)
    sheet_names: list[str] = workbook.sheetnames
    total_changed = 0

    total_cells = _estimate_total_cells(workbook)
    processed_cells = 0

    if on_progress:
        on_progress(0, total_cells)

    def _row_cb(n: int) -> None:
        nonlocal processed_cells
        processed_cells += n
        # on_progress is always truthy here — _row_cb is only passed when
        # on_progress is not None, so the check is redundant and removed.
        on_progress(min(processed_cells, total_cells), total_cells)  # type: ignore[misc]

    for idx, name in enumerate(sheet_names, start=1):
        if cancel_check and cancel_check():
            log.info("Cancellation before sheet '%s'. Stopping.", name)
            break
        if on_sheet_start:
            on_sheet_start(idx, len(sheet_names), name)
        log.info("[%d/%d] Normalising sheet: '%s'", idx, len(sheet_names), name)
        changed = normalize_worksheet(
            workbook[name],
            replace_zwnj=replace_zwnj,
            cancel_check=cancel_check,
            row_callback=_row_cb if on_progress else None,
        )
        log.debug("  -> %d cell(s) modified in '%s'", changed, name)
        total_changed += changed

    log.info("Saving to: %s (atomic write)", destination)
    _atomic_save(workbook, destination)
    log.info("Done. Total cells modified: %d", total_changed)


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------


def build_output_path(source: Path) -> Path:
    """Return *source* with ``_normalized`` appended before the suffix."""
    return source.with_stem(f"{source.stem}_normalized")


def auto_detect_input(directory: Path) -> Path:
    """Return the first .xlsx / .xlsm in *directory* that is not already normalised.

    Raises:
        FileNotFoundError: If no suitable candidate exists.
    """
    candidates = sorted(
        p for p in directory.iterdir()
        if p.suffix.lower() in {".xlsx", ".xlsm"}
        and not p.stem.endswith("_normalized")
    )
    if not candidates:
        raise FileNotFoundError(
            f"No .xlsx / .xlsm files found in '{directory}'. "
            "Provide an explicit INPUT path or change the working directory."
        )
    return candidates[0]


# ---------------------------------------------------------------------------
# CLI layer
# ---------------------------------------------------------------------------


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="normalize_excel",
        description=(
            "Normalise Persian text inside Excel (.xlsx / .xlsm) files.\n\n"
            "Transformations applied:\n"
            "  * Arabic chars  ->  Persian equivalents\n"
            "  * ZWNJ (U+200C) ->  regular space  (unless --keep-zwnj)\n"
            "  * Invisible Unicode control chars removed\n"
            "  * NBSP (U+00A0) ->  regular space\n"
            "  * Multiple spaces -> single space\n"
            "  * Leading/trailing whitespace stripped\n"
            "  * Formula cells (starting with '=') are left untouched"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "input", nargs="?", type=Path, default=None, metavar="INPUT",
        help="Source .xlsx / .xlsm file. Auto-detected in cwd if omitted.",
    )
    parser.add_argument(
        "-o", "--output", type=Path, default=None, metavar="OUTPUT",
        help="Destination path. Defaults to <INPUT_stem>_normalized.xlsx.",
    )
    parser.add_argument(
        "--keep-zwnj", action="store_true", default=False,
        help="Preserve ZWNJ (U+200C) instead of replacing with a space.",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true", default=False,
        help="Enable DEBUG-level logging.",
    )
    return parser


def _configure_cli_logging(*, verbose: bool) -> None:
    """Configure the dedicated package logger for CLI use only."""
    pkg_logger = logging.getLogger("excel_normalizer")
    pkg_logger.setLevel(logging.DEBUG if verbose else logging.INFO)
    if not pkg_logger.handlers:
        handler = logging.StreamHandler(stream=sys.stderr)
        handler.setFormatter(
            logging.Formatter(
                "%(asctime)s [%(levelname)-8s] %(name)s: %(message)s",
                datefmt="%H:%M:%S",
            )
        )
        pkg_logger.addHandler(handler)
    pkg_logger.propagate = False


def main(argv: Optional[list[str]] = None) -> int:
    """CLI entry point. Returns 0 on success, 1 on error."""
    parser = _build_parser()
    args = parser.parse_args(argv)
    _configure_cli_logging(verbose=args.verbose)

    try:
        if args.input is None:
            source = auto_detect_input(Path.cwd())
            log.info("Auto-detected input: %s", source)
        else:
            source = args.input
        destination: Path = args.output or build_output_path(source)
        normalize_workbook(source, destination, replace_zwnj=not args.keep_zwnj)
    except (FileNotFoundError, ValueError) as exc:
        log.error("%s", exc)
        return 1
    except OSError as exc:
        log.error("I/O error: %s", exc)
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
