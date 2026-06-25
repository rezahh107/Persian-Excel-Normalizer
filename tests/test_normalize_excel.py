from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

import normalize_excel
from normalize_excel import (
    NormalizationCancelled,
    _atomic_save,
    build_output_path,
    normalize_text,
    normalize_worksheet,
    normalize_workbook,
)


def _make_workbook(path: Path, value: str = "علي") -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = value
    ws["A2"] = "=A1"
    wb.save(path)


def test_normalize_text_arabic_yeh_and_kaf() -> None:
    assert normalize_text("علي كيان") == "علی کیان"


def test_normalize_text_zwnj_replace_and_preserve() -> None:
    text = "نیم\u200cفاصله"
    assert normalize_text(text) == "نیم فاصله"
    assert normalize_text(text, replace_zwnj=False) == text


def test_normalize_text_invisible_nbsp_and_multi_space() -> None:
    text = "  الف\u200f\xa0\xa0  ب\u200e   ج  "
    assert normalize_text(text) == "الف ب ج"


def test_normalize_text_non_string_passthrough() -> None:
    marker = object()
    assert normalize_text(marker) is marker
    assert normalize_text(123) == 123


def test_normalize_worksheet_preserves_formula_cells() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=SUM(1,2)"
    ws["B1"] = "علي كيان"

    changed = normalize_worksheet(ws)

    assert changed == 1
    assert ws["A1"].value == "=SUM(1,2)"
    assert ws["B1"].value == "علی کیان"


def test_text_starting_with_equals_is_skipped_by_conservative_guard() -> None:
    """Literal text beginning with '=' is intentionally skipped.

    This documents the conservative string-prefix backup in _is_formula_cell().
    Such cells may be real text, but the tool prefers not touching them to
    avoid corrupting formula-like content.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "=برابري"
    ws["A1"].data_type = "s"

    changed = normalize_worksheet(ws)

    assert changed == 0
    assert ws["A1"].value == "=برابري"


def test_xlsx_workbook_smoke_preserves_formula_and_style(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    dst = tmp_path / "nested" / "output.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "علي كيان"
    ws["A2"] = "=A1"
    ws["A1"].font = openpyxl.styles.Font(bold=True)
    wb.save(src)

    normalize_workbook(src, dst)

    out = openpyxl.load_workbook(dst)
    out_ws = out.active
    assert out_ws["A1"].value == "علی کیان"
    assert out_ws["A1"].font.bold is True
    assert out_ws["A2"].value == "=A1"


def test_xlsm_loads_with_keep_vba(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    src = tmp_path / "macro.xlsm"
    dst = tmp_path / "macro_normalized.xlsm"
    src.write_bytes(b"placeholder")
    captured: dict[str, object] = {}

    class DummySheet:
        max_row = 0
        max_column = 0

        def iter_rows(self):
            return iter(())

    class DummyWorkbook:
        worksheets = [DummySheet()]
        sheetnames = ["Sheet"]

        def __getitem__(self, key: str):
            return self.worksheets[0]

    def fake_load_workbook(path: Path, **kwargs):
        captured["path"] = path
        captured.update(kwargs)
        return DummyWorkbook()

    monkeypatch.setattr(openpyxl, "load_workbook", fake_load_workbook)
    monkeypatch.setattr(normalize_excel, "_atomic_save", lambda workbook, destination: None)

    normalize_workbook(src, dst)

    assert captured["path"] == src
    assert captured["keep_vba"] is True


def test_xlsx_loads_without_keep_vba(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    src = tmp_path / "plain.xlsx"
    dst = tmp_path / "plain_normalized.xlsx"
    src.write_bytes(b"placeholder")
    captured: dict[str, object] = {}

    class DummySheet:
        max_row = 0
        max_column = 0

        def iter_rows(self):
            return iter(())

    class DummyWorkbook:
        worksheets = [DummySheet()]
        sheetnames = ["Sheet"]

        def __getitem__(self, key: str):
            return self.worksheets[0]

    def fake_load_workbook(path: Path, **kwargs):
        captured.update(kwargs)
        return DummyWorkbook()

    monkeypatch.setattr(openpyxl, "load_workbook", fake_load_workbook)
    monkeypatch.setattr(normalize_excel, "_atomic_save", lambda workbook, destination: None)

    normalize_workbook(src, dst)

    assert captured["keep_vba"] is False


def test_build_output_path_preserves_suffix() -> None:
    assert build_output_path(Path("data.xlsx")) == Path("data_normalized.xlsx")
    assert build_output_path(Path("macro.xlsm")) == Path("macro_normalized.xlsm")


def test_rejects_same_source_and_destination(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    _make_workbook(src)

    with pytest.raises(ValueError, match="overwrite the input"):
        normalize_workbook(src, src)


def test_rejects_output_without_excel_suffix(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    _make_workbook(src)

    with pytest.raises(ValueError, match="Unsupported output extension"):
        normalize_workbook(src, tmp_path / "output")


def test_rejects_xlsm_to_xlsx_output(tmp_path: Path) -> None:
    src = tmp_path / "macro.xlsm"
    src.write_bytes(b"placeholder")

    with pytest.raises(ValueError, match="Output extension must match"):
        normalize_workbook(src, tmp_path / "macro_normalized.xlsx")


def test_missing_destination_directory_is_created(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    dst = tmp_path / "missing" / "dir" / "output.xlsx"
    _make_workbook(src)

    normalize_workbook(src, dst)

    assert dst.exists()
    out = openpyxl.load_workbook(dst)
    assert out.active["A1"].value == "علی"


def test_cancel_does_not_save_partial_output(tmp_path: Path) -> None:
    src = tmp_path / "input.xlsx"
    dst = tmp_path / "output.xlsx"
    _make_workbook(src, "علي كيان")
    calls = 0

    def cancel_after_first_check() -> bool:
        nonlocal calls
        calls += 1
        return calls > 1

    with pytest.raises(NormalizationCancelled):
        normalize_workbook(src, dst, cancel_check=cancel_after_first_check)

    assert not dst.exists()


def test_atomic_save_failure_cleans_temp_and_preserves_existing(tmp_path: Path) -> None:
    dst = tmp_path / "output.xlsx"
    dst.write_text("original", encoding="utf-8")

    class FailingWorkbook:
        def save(self, path: Path) -> None:
            Path(path).write_text("partial", encoding="utf-8")
            raise RuntimeError("boom")

    with pytest.raises(RuntimeError, match="boom"):
        _atomic_save(FailingWorkbook(), dst)  # type: ignore[arg-type]

    assert dst.read_text(encoding="utf-8") == "original"
    assert list(tmp_path.glob(f".{dst.name}.*")) == []


def test_atomic_save_cleanup_failure_does_not_hide_original(
    monkeypatch: pytest.MonkeyPatch,
    tmp_path: Path,
) -> None:
    dst = tmp_path / "output.xlsx"

    class FailingWorkbook:
        def save(self, path: Path) -> None:
            Path(path).write_text("partial", encoding="utf-8")
            raise RuntimeError("original error")

    def fail_unlink(self: Path, missing_ok: bool = False) -> None:
        raise PermissionError("cleanup failed")

    monkeypatch.setattr(Path, "unlink", fail_unlink)

    with pytest.raises(RuntimeError, match="original error"):
        _atomic_save(FailingWorkbook(), dst)  # type: ignore[arg-type]
